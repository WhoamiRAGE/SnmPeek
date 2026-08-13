"""Excel export: dumps the current device table and full status history
to a .xlsx workbook via openpyxl.
"""

from __future__ import annotations

from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.engine import Engine

from storage.db import devices_table, status_history_table

_DEVICE_HEADERS = ["IP", "MAC", "Vendor", "Hostname", "SNMP Enabled", "Status", "First Seen", "Last Seen"]
_HISTORY_HEADERS = ["IP", "Status", "Timestamp"]


def _autosize(ws, headers: list[str]) -> None:
    for i, header in enumerate(headers, start=1):
        col = get_column_letter(i)
        lengths = [len(header)] + [len(str(c.value)) for c in ws[col][1:] if c.value is not None]
        ws.column_dimensions[col].width = min(max(lengths) + 2, 60)


def export_to_excel(engine: Engine, output_path: str) -> str:
    """Write devices + status_history tables to an .xlsx file. Returns the path."""
    wb = Workbook()

    devices_ws = wb.active
    devices_ws.title = "Devices"
    devices_ws.append(_DEVICE_HEADERS)
    for cell in devices_ws[1]:
        cell.font = Font(bold=True)

    with engine.connect() as conn:
        for row in conn.execute(select(devices_table)):
            devices_ws.append(
                [
                    row.ip,
                    row.mac,
                    row.vendor,
                    row.hostname,
                    "yes" if row.snmp_enabled else "no",
                    row.status,
                    row.first_seen.strftime("%Y-%m-%d %H:%M:%S") if row.first_seen else None,
                    row.last_seen.strftime("%Y-%m-%d %H:%M:%S") if row.last_seen else None,
                ]
            )
    _autosize(devices_ws, _DEVICE_HEADERS)

    history_ws = wb.create_sheet("Status History")
    history_ws.append(_HISTORY_HEADERS)
    for cell in history_ws[1]:
        cell.font = Font(bold=True)

    with engine.connect() as conn:
        for row in conn.execute(select(status_history_table).order_by(status_history_table.c.timestamp)):
            history_ws.append(
                [
                    row.ip,
                    row.status,
                    row.timestamp.strftime("%Y-%m-%d %H:%M:%S") if row.timestamp else None,
                ]
            )
    _autosize(history_ws, _HISTORY_HEADERS)

    wb.save(output_path)
    return output_path


def default_export_path() -> str:
    return f"snmpeek_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
